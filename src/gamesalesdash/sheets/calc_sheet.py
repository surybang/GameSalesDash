"""Onglet calc_sheet : agrégats par année, genre et éditeur (source de TDB_2)."""

from openpyxl import Workbook

from gamesalesdash.config import (
    PUBLISHERS,
    SHEET_CALC,
    SHEET_CLEANED,
    SHEET_RESOURCES,
)


def build_calc_sheet(wb: Workbook) -> None:
    """Construit l'onglet calc_sheet contenant les agrégats intermédiaires.

    Ces tableaux servent de source aux graphiques de TDB_2 : ventes et nombre
    de jeux par année, par genre, et par éditeur sélectionné.

    Args:
        wb: Classeur cible.
    """
    ws = wb.create_sheet(SHEET_CALC)

    # Années et agrégats par année
    ws["A1"] = "Année"
    ws["B1"] = "Ventes"
    ws["C1"] = "Nombre de jeux"
    for i in range(2, 40):
        ws.cell(row=i, column=1).value = f"={SHEET_RESOURCES}!B{i}"
        ws.cell(row=i, column=2).value = f"=SUMIFS({SHEET_CLEANED}!K:K,{SHEET_CLEANED}!D:D,A{i})"
        ws.cell(row=i, column=3).value = f'=COUNTIFS({SHEET_CLEANED}!B:B,"<>",{SHEET_CLEANED}!D:D,A{i})'

    # Genres et agrégats par genre
    ws["E1"] = "Nombre de jeux"
    ws["F1"] = "Total ventes"
    for i in range(1, 14):
        ws.cell(row=i, column=4).value = f"={SHEET_RESOURCES}!C{i}"
    for i in range(2, 14):
        ws.cell(row=i, column=5).value = f'=COUNTIFS({SHEET_CLEANED}!B:B,"<>",{SHEET_CLEANED}!E:E,D{i})'
        ws.cell(row=i, column=6).value = f"=SUMIFS({SHEET_CLEANED}!K:K,{SHEET_CLEANED}!E:E,D{i})"

    # Éditeurs sélectionnés et agrégats
    ws["G1"] = "Editeurs"
    ws["H1"] = "Nombre de jeux"
    ws["I1"] = "Total ventes"
    for i, publisher in enumerate(PUBLISHERS, start=2):
        ws.cell(row=i, column=7).value = publisher
        ws.cell(row=i, column=8).value = f'=COUNTIFS({SHEET_CLEANED}!F:F,G{i},{SHEET_CLEANED}!B:B,"<>")'
        ws.cell(row=i, column=9).value = f"=SUMIFS({SHEET_CLEANED}!K:K,{SHEET_CLEANED}!F:F,G{i})"
