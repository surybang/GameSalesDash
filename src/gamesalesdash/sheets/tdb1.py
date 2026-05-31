"""Onglet TDB_1 : tableau de bord principal (filtres, TOP 5, graphiques, indicateurs)."""

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from gamesalesdash.components.charts import create_bar_chart
from gamesalesdash.config import (
    DARK_BLUE,
    DARK_GREEN,
    LIGHT_BLUE,
    LIGHT_GREEN,
    SHEET_RESOURCES,
    SHEET_TDB1,
)
from gamesalesdash.components.filters import add_filter
from gamesalesdash.components.indicators import add_title_map, add_value_map
from gamesalesdash.components.styles import apply_border_style
from gamesalesdash.components.tables import add_column_names_tdb, add_title_tdb, apply_formulas


def build_tdb1(wb: Workbook, len_dict: dict[str, int]) -> None:
    """Construit le tableau de bord principal TDB_1.

    Assemble dans l'ordre les filtres, les tableaux TOP 5, les graphiques,
    les cartes indicateurs puis ajuste les largeurs de colonnes.

    Args:
        wb: Classeur cible.
        len_dict: Tailles des listes uniques pour les listes de validation.
    """
    ws = wb.create_sheet(SHEET_TDB1)
    ws.sheet_view.showGridLines = False

    _add_filters(ws, len_dict)
    _add_tables(ws)
    _add_charts(ws)
    _add_indicators(ws)


def _add_filters(ws: Worksheet, len_dict: dict[str, int]) -> None:
    """Ajoute les filtres déroulants Année, Plateforme et Genre.

    Args:
        ws: Feuille TDB_1.
        len_dict: Tailles des listes uniques pour borner les plages de validation.
    """
    add_filter(ws, "Année", 2009, 1, 1, 2, 2,
               f"='{SHEET_RESOURCES}'!$B$1:$B${len_dict['len_year']}")
    add_filter(ws, "Plateforme", "PS3", 1, 5, 2, 6,
               f"='{SHEET_RESOURCES}'!$A$1:$A${len_dict['len_platform']}")
    add_filter(ws, "Genre", "Shooter", 1, 9, 2, 10,
               f"='{SHEET_RESOURCES}'!$C$1:$C${len_dict['len_genre']}")
    apply_border_style(ws, 1, 2, 1, 12, 2)


def _add_tables(ws: Worksheet) -> None:
    """Ajoute les titres, en-têtes et formules des quatre tableaux TOP 5.

    Args:
        ws: Feuille TDB_1.
    """
    # Titres dynamiques
    add_title_tdb(ws, "A5", "G1", DARK_BLUE, 5, 1, 5, 4)
    add_title_tdb(ws, "F5", "G1", DARK_BLUE, 5, 6, 5, 9, filter_genre="K1")
    add_title_tdb(ws, "A15", "G3", DARK_GREEN, 15, 1, 15, 4)
    add_title_tdb(ws, "F15", "G3", DARK_GREEN, 15, 6, 15, 9, filter_genre="K1")

    # En-têtes de colonnes
    add_column_names_tdb(ws, 6, 1, 4, LIGHT_BLUE)
    add_column_names_tdb(ws, 6, 6, 9, LIGHT_BLUE)
    add_column_names_tdb(ws, 16, 1, 4, LIGHT_GREEN)
    add_column_names_tdb(ws, 16, 6, 9, LIGHT_GREEN)

    # Formules TOP 5: B=Nom, K=Ventes globales, G=Ventes NA, H=Ventes EU
    targets = ["B", "K", "G", "H"]
    tops = [1, 2, 3, 4, 5]
    apply_formulas(ws, ("A", 7), tops, targets, "$G$1")
    apply_formulas(ws, ("A", 17), tops, targets, "$G$3")
    apply_formulas(ws, ("F", 7), tops, targets, "$G$1", genre="$K$1")
    apply_formulas(ws, ("F", 17), tops, targets, "$G$3", genre="$K$1")


def _add_charts(ws: Worksheet) -> None:
    """Ajoute les deux graphiques à barres des TOP 5 par plateforme.

    Args:
        ws: Feuille TDB_1.
    """
    platform1 = ws["G1"].value
    platform2 = ws["G3"].value

    create_bar_chart(
        ws, f"Top 5 des jeux vendus {platform1}",
        {"min_col": 2, "min_row": 6, "max_row": 11, "max_col": 4},
        {"min_col": 1, "min_row": 7, "max_row": 11},
        "A23", style=11,
    )
    create_bar_chart(
        ws, f"Top 5 des jeux vendus {platform2}",
        {"min_col": 2, "min_row": 16, "max_row": 21, "max_col": 4},
        {"min_col": 1, "min_row": 17, "max_row": 21},
        "F23", style=13,
    )


def _add_indicators(ws: Worksheet) -> None:
    """Ajoute les cartes indicateurs (titres et valeurs COUNTIFS / SUMIFS).

    Args:
        ws: Feuille TDB_1.
    """
    # (row, col, platform, color, genre, type_value)
    title_configs = [
        (5, 12, "G1", DARK_BLUE, "", "count"),
        (5, 13, "G1", DARK_BLUE, "K1", "count"),
        (9, 12, "G1", DARK_BLUE, "", "sum"),
        (9, 13, "G1", DARK_BLUE, "K1", "sum"),
        (15, 12, "G3", DARK_GREEN, "", "sum"),
        (15, 13, "G3", DARK_GREEN, "K1", "sum"),
        (19, 12, "G3", DARK_GREEN, "", "sum"),
        (19, 13, "G3", DARK_GREEN, "K1", "sum"),
    ]
    value_configs = [
        (7, 12, "G1", LIGHT_BLUE, "", "count"),
        (7, 13, "G1", LIGHT_BLUE, "K1", "count"),
        (11, 12, "G1", LIGHT_BLUE, "", "sum"),
        (11, 13, "G1", LIGHT_BLUE, "K1", "sum"),
        (17, 12, "G3", LIGHT_GREEN, "", "count"),
        (17, 13, "G3", LIGHT_GREEN, "K1", "count"),
        (21, 12, "G3", LIGHT_GREEN, "", "sum"),
        (21, 13, "G3", LIGHT_GREEN, "K1", "sum"),
    ]
    for row, col, platform, color, genre, type_val in title_configs:
        add_title_map(ws, f"{get_column_letter(col)}{row}", platform, color,
                      row, col, row + 1, col, type_val, filter_genre=genre)

    for row, col, platform, color, genre, type_val in value_configs:
        add_value_map(ws, f"{get_column_letter(col)}{row}", platform, color,
                      row, col, row + 1, col, type_val, filter_genre=genre)
