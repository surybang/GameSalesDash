"""Onglet Ressources : listes de valeurs uniques pour les filtres."""

from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

from gamesalesdash.config import (
    COL_GENRE,
    COL_PLATFORM,
    COL_PUBLISHER,
    COL_YEAR,
    SHEET_CLEANED,
    SHEET_RESOURCES,
)


def build_resources_sheet(wb: Workbook, len_dict: dict[str, int]) -> None:
    """Crée l'onglet Ressources et y insère les listes de valeurs uniques.

    Les formules UNIQUE et SORT ne sont pas reconnues nativement par openpyxl,
    elles sont préfixées avec _xlfn. pour être interprétées par Excel.

    Structure de l'onglet :
        - Colonne A : plateformes uniques
        - Colonne B : années uniques triées
        - Colonne C : genres uniques
        - Colonne D : éditeurs uniques

    Args:
        wb: Classeur dans lequel créer l'onglet.
        len_dict: Tailles des listes uniques (clés len_platform, len_year,
            len_genre, len_publisher), utilisées pour borner les plages des
            formules matricielles.
    """
    ws = wb.create_sheet(SHEET_RESOURCES)
    ref = SHEET_CLEANED

    # Plateformes (col A)
    formula = f"=_xlfn.UNIQUE({ref}!{COL_PLATFORM}:{COL_PLATFORM})"
    ws["A1"] = ArrayFormula(f"A1:A{len_dict['len_platform']}", formula)

    # Années triées (col B)
    formula = f"=_xlfn.SORT(_xlfn.UNIQUE({ref}!{COL_YEAR}:{COL_YEAR}))"
    ws["B1"] = ArrayFormula(f"B1:B{len_dict['len_year']}", formula)

    # Genres (col C)
    formula = f"=_xlfn.UNIQUE({ref}!{COL_GENRE}:{COL_GENRE})"
    ws["C1"] = ArrayFormula(f"C1:C{len_dict['len_genre']}", formula)

    # Éditeurs (col D)
    formula = f"=_xlfn.UNIQUE({ref}!{COL_PUBLISHER}:{COL_PUBLISHER})"
    ws["D1"] = ArrayFormula(f"D1:D{len_dict['len_publisher']}", formula)
