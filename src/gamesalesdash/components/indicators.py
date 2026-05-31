"""Cartes indicateurs : titres dynamiques et formules COUNTIFS / SUMIFS."""

from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from gamesalesdash.config import SHEET_CLEANED


def add_title_map(
    sheet: Worksheet,
    cell: str,
    filter_platform: str,
    color: str,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
    type_value: str,
    filter_year: str = "C1",
    filter_genre: str = "",
) -> None:
    """Insère le titre dynamique d'une carte indicateur.

    Le titre est une formule Excel concaténant les valeurs des filtres, dont
    le libellé varie selon le type d'indicateur et la présence d'un filtre genre.

    Args:
        sheet: Feuille Excel cible.
        cell: Cellule d'ancrage du titre (ex : "L5").
        filter_platform: Référence de la cellule de filtre plateforme (ex : "G1").
        color: Couleur de fond hex sans # (ex : "8DB4E2").
        start_row: Première ligne de la plage à fusionner.
        start_column: Première colonne de la plage à fusionner.
        end_row: Dernière ligne de la plage à fusionner.
        end_column: Dernière colonne de la plage à fusionner.
        type_value: "count" pour un dénombrement, "sum" pour un total de ventes.
        filter_year: Référence de la cellule de filtre année (défaut : "C1").
        filter_genre: Référence de la cellule de filtre genre. Chaîne vide si absent.

    Raises:
        ValueError: Si la combinaison type_value / filter_genre n'est pas gérée.
    """
    title_cell = sheet[cell]

    if filter_genre and type_value == "count":
        formula = (
            f'="Nombre de jeux de type " & {filter_genre}'
            f' & " sur " & {filter_platform} & " en " & {filter_year}'
        )
    elif not filter_genre and type_value == "count":
        formula = (
            f'="Nombre de jeux sortis sur " & {filter_platform}'
            f' & " en " & {filter_year}'
        )
    elif filter_genre and type_value == "sum":
        formula = (
            f'="Nombre de jeux de type " & {filter_genre}'
            f' & " vendus sur " & {filter_platform} & " en " & {filter_year}'
        )
    elif not filter_genre and type_value == "sum":
        formula = (
            f'="Nombre de jeux vendus sur " & {filter_platform}'
            f' & " en " & {filter_year}'
        )
    else:
        raise ValueError(
            f"Combinaison non gérée : type_value={type_value!r}, "
            f"filter_genre={filter_genre!r}"
        )

    title_cell.value = formula
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column,
    )


def add_value_map(
    sheet: Worksheet,
    cell: str,
    filter_platform: str,
    color: str,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
    type_value: str,
    filter_year: str = "C1",
    filter_genre: str = "",
) -> None:
    """Insère la formule de calcul d'une carte indicateur (COUNTIFS ou SUMIFS).

    Args:
        sheet: Feuille Excel cible.
        cell: Cellule d'ancrage de la valeur (ex : "L7").
        filter_platform: Référence de la cellule de filtre plateforme (ex : "G1").
        color: Couleur de fond hex sans # (ex : "C5D9F1").
        start_row: Première ligne de la plage à fusionner.
        start_column: Première colonne de la plage à fusionner.
        end_row: Dernière ligne de la plage à fusionner.
        end_column: Dernière colonne de la plage à fusionner.
        type_value: "count" pour COUNTIFS, "sum" pour SUMIFS sur les ventes EU.
        filter_year: Référence de la cellule de filtre année (défaut : "C1").
        filter_genre: Référence de la cellule de filtre genre. Chaîne vide si absent.

    Raises:
        ValueError: Si type_value n'est ni "count" ni "sum".
    """
    title_cell = sheet[cell]
    base = f"{SHEET_CLEANED}!$D:$D,{filter_year},{SHEET_CLEANED}!$C:$C,{filter_platform}"

    if type_value == "count":
        formula = (
            f"=COUNTIFS({base},{SHEET_CLEANED}!$E:$E,{filter_genre})"
            if filter_genre
            else f"=COUNTIFS({base})"
        )
    elif type_value == "sum":
        sum_range = f"{SHEET_CLEANED}!$H:$H"
        formula = (
            f'=SUMIFS({sum_range},{base},{SHEET_CLEANED}!$E:$E,{filter_genre}) & "M"'
            if filter_genre
            else f'=SUMIFS({sum_range},{base}) & "M"'
        )
    else:
        raise ValueError(f"type_value non géré : {type_value!r}")

    title_cell.value = formula
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column,
    )
