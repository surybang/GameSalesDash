"""Tableaux TOP 5 : titres dynamiques, en-têtes de colonnes et formules matricielles."""

from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from gamesalesdash.config import SHEET_CLEANED


def add_title_tdb(
    sheet: Worksheet,
    cell: str,
    filter_platform: str,
    color: str,
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
    filter_year: str = "C1",
    filter_genre: str | None = None,
) -> None:
    """Insère un titre dynamique pour un tableau TOP 5.

    Le titre est une formule Excel qui concatène les valeurs des cellules de
    filtre, il se met donc à jour automatiquement quand l'utilisateur change
    un filtre.

    Args:
        sheet: Feuille Excel cible.
        cell: Cellule d'ancrage du titre (ex : "A5").
        filter_platform: Référence de la cellule de filtre plateforme (ex : "G1").
        color: Couleur de fond hex sans # (ex : "8DB4E2").
        start_row: Première ligne de la plage à fusionner.
        start_column: Première colonne de la plage à fusionner.
        end_row: Dernière ligne de la plage à fusionner.
        end_column: Dernière colonne de la plage à fusionner.
        filter_year: Référence de la cellule de filtre année (défaut : "C1").
        filter_genre: Référence de la cellule de filtre genre. Si None, le titre
            n'inclut pas le genre.
    """
    title_cell = sheet[cell]

    if filter_genre:
        formula = (
            f'="Les meilleurs " & {filter_genre}'
            f' & " sur " & {filter_platform}'
            f' & " en " & {filter_year}'
        )
    else:
        formula = (
            f'="Les meilleurs jeux sur " & {filter_platform}'
            f' & " en " & {filter_year}'
        )

    title_cell.value = formula
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=end_row,
        end_column=end_column,
    )


def add_column_names_tdb(
    sheet: Worksheet,
    start_row: int,
    start_column: int,
    end_column: int,
    color: str,
) -> None:
    """Insère les en-têtes de colonnes d'un tableau TOP 5 avec mise en forme.

    Args:
        sheet: Feuille Excel cible.
        start_row: Ligne où écrire les en-têtes.
        start_column: Première colonne (correspond à "Nom").
        end_column: Dernière colonne (correspond à "Ventes EU").
        color: Couleur de fond hex sans # (ex : "C5D9F1").
    """
    column_names = ["Nom", "Ventes Totales", "Ventes NA", "Ventes EU"]
    for col_num, col_name in zip(range(start_column, end_column + 1), column_names):
        cell = sheet.cell(row=start_row, column=col_num)
        cell.value = col_name
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def apply_formula(
    sheet: Worksheet,
    cellule: str,
    top: int,
    platform: str,
    target: str,
    year: str = "$C$1",
    ref_sheet: str = SHEET_CLEANED,
    genre: str | None = None,
) -> None:
    """Applique une formule matricielle INDEX/MATCH/LARGE pour récupérer le Nième jeu.

    La formule filtre cleaned_data sur plateforme et année (et genre si fourni),
    puis retourne la valeur de la colonne cible correspondant au Nième meilleur
    score de ventes globales.

    Args:
        sheet: Feuille Excel cible.
        cellule: Adresse de la cellule recevant la formule (ex : "A7").
        top: Rang à retourner (1 = meilleur, 2 = second, etc.).
        platform: Référence de la cellule de filtre plateforme (ex : "$G$1").
        target: Lettre de la colonne à retourner depuis cleaned_data
            (ex : "B" pour le nom).
        year: Référence de la cellule de filtre année (défaut : "$C$1").
        ref_sheet: Nom de l'onglet source des données.
        genre: Référence de la cellule de filtre genre. Si None, le filtre
            genre n'est pas appliqué.
    """
    if genre:
        formula = (
            f"=INDEX({ref_sheet}!{target}:{target},"
            f"MATCH("
            f"LARGE(IF(({ref_sheet}!$C:$C={platform})"
            f"*({ref_sheet}!$D:$D={year})"
            f"*({ref_sheet}!$E:$E={genre}),"
            f"{ref_sheet}!$K:$K),{top}),"
            f"IF(({ref_sheet}!$C:$C={platform})"
            f"*({ref_sheet}!$D:$D={year})"
            f"*({ref_sheet}!$E:$E={genre}),"
            f"{ref_sheet}!$K:$K),0))"
        )
    else:
        formula = (
            f"=INDEX({ref_sheet}!{target}:{target},"
            f"MATCH("
            f"LARGE(IF(({ref_sheet}!$C:$C={platform})"
            f"*({ref_sheet}!$D:$D={year}),"
            f"{ref_sheet}!$K:$K),{top}),"
            f"IF(({ref_sheet}!$C:$C={platform})"
            f"*({ref_sheet}!$D:$D={year}),"
            f"{ref_sheet}!$K:$K),0))"
        )

    sheet[cellule] = ArrayFormula(f"{cellule}:{cellule}", formula)


def apply_formulas(
    sheet: Worksheet,
    start_cell: tuple[str, int],
    top_values: list[int],
    targets: list[str],
    platform: str,
    genre: str | None = None,
) -> None:
    """Applique les formules TOP N sur toutes les cellules d'un tableau.

    Boucle sur les rangs (lignes) et les colonnes cibles pour remplir le
    tableau complet à partir de sa cellule de départ.

    Args:
        sheet: Feuille Excel cible.
        start_cell: Tuple (lettre de colonne, numéro de ligne) de la première
            cellule du tableau (ex : ("A", 7)).
        top_values: Liste des rangs à afficher (ex : [1, 2, 3, 4, 5]).
        targets: Colonnes de cleaned_data à retourner, dans l'ordre des colonnes
            du tableau (ex : ["B", "K", "G", "H"] pour nom, global, NA, EU).
        platform: Référence de la cellule de filtre plateforme.
        genre: Référence de la cellule de filtre genre, ou None.
    """
    for i, top in enumerate(top_values):
        for j, target in enumerate(targets):
            col_letter = chr(ord(start_cell[0]) + j)
            row = start_cell[1] + i
            apply_formula(
                sheet=sheet,
                cellule=f"{col_letter}{row}",
                top=top,
                target=target,
                platform=platform,
                genre=genre,
            )
